# Importa/Actualiza datos Excel -> Tabla

import openpyxl
from flask import Flask, flash
import datetime as dt

class Importa:

    def __init__(self, cx):
        self.cx = cx
        self.cur = cx.cursor()

    def importa_dist(self, excel_file, table_name, dep, usr):
        ''' Actualiza DIST a partir de archivo excel '''
        # 1ro elimina dist de departamento
        del_query = f'''
            DELETE FROM [GeografiaElectoral_app].[dbo].[{table_name}] WHERE IdLocDist
            IN (SELECT IdLoc from [GeografiaElectoral_app].[dbo].[LOC] 
                WHERE DepLoc = {dep})
        '''
        self.cur.execute(del_query)

        c = 0
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active     #Seleccionar la hoja activa
        fecha = dt.datetime.now()

        for row in sheet.iter_rows(min_row=2, values_only=True): # 2 para omitir el encabezado

            # verifica si existe
            verif_query =  f'select * from [GeografiaElectoral_app].[dbo].LOC ' \
                    ' where idloc = %s and deploc = %s'

            verif = row[0], dep
            self.cur.execute(verif_query, verif)
            existe = self.cur.fetchone()
            if not existe:
                flash(f'IdLocReci: {row[0]}  Asiento No encontrado en departamento seleccionado - ({dep}) !!! ...debe corregir este dato...', 'alert-info')
                return False

            # Inserta
            insert_query = f'insert into  [GeografiaElectoral_app].[dbo].[{table_name}] (IdLocDist, Dist, CircunDist, NomDist, fechaIngreso, fechaAct, usuario) ' \
                    ' values (%s, %s, %s, %s, %s, %s, %s) '

            t = row[0], row[1], row[2], row[3], fecha, fecha, usr  # Cambia según el número de columnas 
            try:
                self.cur.execute(insert_query, t)
                c += 1
            except Exception as e:
                flash(e, 'alert-info')
                return False

        # Confirmar cambios y cerrar la conexión
        self.cx.commit()
        self.cur.close()
        flash(f'Proceso concluido !! ... cantidad de registros importados: {c}', 'alert-success')
        return True


    def importa_zona(self, excel_file, table_name, dep, usr):
        ''' Actualiza ZONA a partir de archivo excel '''

        # 1ro elimina zona de departamento
        del_query = f'''
            DELETE FROM [GeografiaElectoral_app].[dbo].[{table_name}] WHERE IdLocZona
            IN (SELECT IdLoc from [GeografiaElectoral_app].[dbo].[LOC] 
                WHERE DepLoc = {dep})
        '''
        self.cur.execute(del_query)

        c = 0
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active     #Seleccionar la hoja activa
        fecha = dt.datetime.now()

        for row in sheet.iter_rows(min_row=2, values_only=True): # 2 para omitir el encabezado

            # verifica si existe
            verif_query =  f'select * from [GeografiaElectoral_app].[dbo].LOC ' \
                    ' where idloc = %s and deploc = %s'

            verif = row[0], dep
            self.cur.execute(verif_query, verif)
            existe = self.cur.fetchone()
            if not existe:
                flash(f'IdLocReci: {row[0]}  Asiento No encontrado en departamento seleccionado - ({dep}) !!! ...debe corregir este dato...', 'alert-info')
                return False

            # Inserta
            insert_query = f'insert into  [GeografiaElectoral_app].[dbo].[{table_name}] (IdLocZona, Zona, NomZona, DistZona, fechaIngreso, fechaAct, usuario) ' \
                    ' values (%s, %s, %s, %s, %s, %s, %s) '

            t = row[0], row[1], row[2], row[3], fecha, fecha, usr  # Cambia según el número de columnas 
            try:
                self.cur.execute(insert_query, t)
                c += 1
            except Exception as e:
                flash(e, 'alert-info')
                return False

        # Confirmar cambios y cerrar la conexión
        self.cx.commit()
        self.cur.close()
        flash(f'Proceso concluido !! ... cantidad de registros importados: {c}', 'alert-success')
        return True


    def upd_zona_reci(self, excel_file, table_name, dep, usr):
        ''' actualiza ZONA de tabla RECI a partir de archivo excel '''

        c = 0
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active     #Seleccionar la hoja activa
        fecha = dt.datetime.now()

        for row in sheet.iter_rows(min_row=2, values_only=True): # 2 para omitir el encabezado
            # verifica si existe?
            verif_query =  f'select * from [GeografiaElectoral_app].[dbo].[{table_name}] ' \
                    ' where idlocreci = %s and reci =  %s '
            verif = row[0], row[1]
            self.cur.execute(verif_query, verif)
            existe = self.cur.fetchone()
            if not existe:
                flash(f'IdLoc: {row[0]}, Reci: {row[1]}  No encontrado !!! ...debe corregir estos datos...', 'alert-info')
                return False

            # actualiza
            upd_query = f'update [GeografiaElectoral_app].[dbo].[{table_name}] ' \
                    ' set zonaReci = %s, fechaAct = %s, usuario = %s ' \
                    ' where idlocreci = %s and reci =  %s '
            t = row[2], fecha, usr, row[0], row[1]  # Cambia según el número de columnas 
            try:
                self.cur.execute(upd_query, t)
                c += 1
            except Exception as e:
                flash(e, 'alert-info')
                return False


        # Confirmar cambios y cerrar la conexión
        self.cx.commit()
        self.cur.close()
        flash(f'Proceso concluido !! ... cantidad de registros actualizados: {c}', 'alert-success')
        return True
